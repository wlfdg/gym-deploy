from flask import Flask, request, jsonify, make_response
from flask_cors import CORS
import psycopg2
import psycopg2.extras
from datetime import datetime, timedelta
import csv
import io
import hashlib
import os

@app.route("/")
def index():
    return jsonify({"status": "Gym API is running"})
app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})
@app.route("/")
def home():
    return jsonify({
        "status": "Gym API running",
        "message": "Backend is working"
    })

# ── Database ──────────────────────────────────────────────────────────────────
# Set DATABASE_URL in Vercel environment variables
# Format: postgresql://postgres:[PASSWORD]@db.[PROJECT-REF].supabase.co:5432/postgres

DATABASE_URL = os.environ.get("DATABASE_URL")

def get_db():
    if not DATABASE_URL:
        raise Exception("DATABASE_URL environment variable not set")

    conn = psycopg2.connect(
        DATABASE_URL,
        sslmode="require"
    )
    return conn

def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

# ── Helpers ────────────────────────────────────────────────────────────────────

def error(msg, code=400):
    return jsonify({"message": msg}), code

def require_fields(data, *fields):
    for f in fields:
        if not data.get(f) and data.get(f) != 0:
            return f"Missing required field: {f}"
    return None

def row_to_dict(row, cursor):
    """Convert psycopg2 row to dict using cursor description."""
    if row is None:
        return None
    columns = [desc[0] for desc in cursor.description]
    return dict(zip(columns, row))

def rows_to_list(rows, cursor):
    columns = [desc[0] for desc in cursor.description]
    return [dict(zip(columns, row)) for row in rows]

# ── Auth ──────────────────────────────────────────────────────────────────────

@app.route("/login", methods=["POST"])
def login():
    data = request.json or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    if not username or not password:
        return error("Username and password are required.")
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM admins WHERE username=%s AND password=%s",
        (username, hash_password(password))
    )
    admin = cur.fetchone()
    cur.close()
    conn.close()
    
    # 👇 This must have no extra spaces or stray text
    if admin:
        return jsonify({"success": True, "username": admin[0] if not hasattr(admin, 'keys') else admin["username"]})
    
    return error("Invalid username or password.", 401)

@app.route("/register", methods=["POST"])
def register():
    data = request.json or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    if not username or not password:
        return error("Username and password are required.")
    if len(password) < 6:
        return error("Password must be at least 6 characters.")
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO admins (username, password) VALUES (%s, %s)",
            (username, hash_password(password))
        )
        conn.commit()
        cur.close()
        conn.close()
        return jsonify({"message": "Admin created"}), 201
    except psycopg2.errors.UniqueViolation:
        return error("Username already exists.", 409)

# ── Members ───────────────────────────────────────────────────────────────────

@app.route("/members", methods=["GET"])
def get_members():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM members ORDER BY id DESC")
    rows = rows_to_list(cur.fetchall(), cur)
    cur.close()
    conn.close()
    return jsonify(rows)

@app.route("/members/<int:member_id>", methods=["GET"])
def get_member(member_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM members WHERE id=%s", (member_id,))
    row = row_to_dict(cur.fetchone(), cur)
    cur.close()
    conn.close()
    if row:
        return jsonify(row)
    return error("Member not found.", 404)

@app.route("/members", methods=["POST"])
def add_member():
    data = request.json or {}
    err = require_fields(data, "name", "months", "price")
    if err:
        return error(err)
    try:
        months = int(data["months"])
        price = float(data["price"])
        discount = float(data.get("discount", 0))
        if months < 1 or months > 60:
            return error("Months must be between 1 and 60.")
        if price < 0:
            return error("Price cannot be negative.")
        if not (0 <= discount <= 100):
            return error("Discount must be between 0 and 100.")
    except (ValueError, TypeError):
        return error("Invalid numeric values.")

    start = datetime.now()
    expiration = start + timedelta(days=30 * months)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO members (name, email, phone, plan, months, price, discount, start_date, expiration_date, status)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'active')
    """, (
        data["name"].strip(),
        data.get("email", "").strip(),
        data.get("phone", "").strip(),
        data.get("plan", "Basic"),
        months, price, discount,
        start.strftime("%Y-%m-%d"),
        expiration.strftime("%Y-%m-%d")
    ))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"message": "Member added"}), 201

@app.route("/members/<int:member_id>", methods=["PUT"])
def update_member(member_id):
    data = request.json or {}
    err = require_fields(data, "name", "months", "price")
    if err:
        return error(err)
    try:
        months = int(data["months"])
        price = float(data["price"])
        discount = float(data.get("discount", 0))
    except (ValueError, TypeError):
        return error("Invalid numeric values.")

    start_str = data.get("start_date") or datetime.now().strftime("%Y-%m-%d")
    try:
        expiration = (datetime.strptime(start_str, "%Y-%m-%d") + timedelta(days=30 * months)).strftime("%Y-%m-%d")
    except ValueError:
        return error("Invalid start_date format. Use YYYY-MM-DD.")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE members SET name=%s, email=%s, phone=%s, plan=%s, months=%s, price=%s,
        discount=%s, start_date=%s, expiration_date=%s WHERE id=%s
    """, (
        data["name"].strip(),
        data.get("email", "").strip(),
        data.get("phone", "").strip(),
        data.get("plan", "Basic"),
        months, price, discount,
        start_str, expiration, member_id
    ))
    conn.commit()
    rows_affected = cur.rowcount
    cur.close()
    conn.close()
    if rows_affected == 0:
        return error("Member not found.", 404)
    return jsonify({"message": "Member updated"})

@app.route("/members/<int:member_id>", methods=["DELETE"])
def delete_member(member_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM members WHERE id=%s", (member_id,))
    conn.commit()
    rows_affected = cur.rowcount
    cur.close()
    conn.close()
    if rows_affected == 0:
        return error("Member not found.", 404)
    return jsonify({"message": "Deleted"})

# ── Attendance ────────────────────────────────────────────────────────────────

@app.route("/attendance/timein", methods=["POST"])
def time_in():
    data = request.json or {}
    member_id = data.get("member_id")
    if not member_id:
        return error("member_id is required.")
    now = datetime.now()
    today = now.strftime("%Y-%m-%d")
    time_now = now.strftime("%I:%M %p")
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT id FROM attendance WHERE member_id=%s AND date=%s AND time_out IS NULL",
        (member_id, today)
    )
    if cur.fetchone():
        cur.close()
        conn.close()
        return error("Member is already timed in today.", 409)
    cur.execute("SELECT name FROM members WHERE id=%s", (member_id,))
    member = cur.fetchone()
    if not member:
        cur.close()
        conn.close()
        return error("Member not found.", 404)
    cur.execute(
        "INSERT INTO attendance (member_id, member_name, time_in, date) VALUES (%s, %s, %s, %s)",
        (member_id, member[0], time_now, today)
    )
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"message": f"{member[0]} timed in at {time_now}"}), 201

@app.route("/attendance/timeout", methods=["POST"])
def time_out():
    data = request.json or {}
    member_id = data.get("member_id")
    if not member_id:
        return error("member_id is required.")
    now = datetime.now()
    today = now.strftime("%Y-%m-%d")
    time_now = now.strftime("%I:%M %p")
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT id, member_name FROM attendance WHERE member_id=%s AND date=%s AND time_out IS NULL",
        (member_id, today)
    )
    record = cur.fetchone()
    if not record:
        cur.close()
        conn.close()
        return error("No active time-in found for this member today.", 404)
    cur.execute("UPDATE attendance SET time_out=%s WHERE id=%s", (time_now, record[0]))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"message": f"{record[1]} timed out at {time_now}"})

@app.route("/attendance", methods=["GET"])
def get_attendance():
    date = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM attendance WHERE date=%s ORDER BY id DESC", (date,))
    rows = rows_to_list(cur.fetchall(), cur)
    cur.execute(
        "SELECT COUNT(*) FROM attendance WHERE date=%s AND time_out IS NULL", (date,)
    )
    still_in = cur.fetchone()[0]
    cur.close()
    conn.close()
    return jsonify({
        "records": rows,
        "total": len(rows),
        "still_in": still_in,
        "date": date
    })

# ── Walk-ins ──────────────────────────────────────────────────────────────────

@app.route("/walkins", methods=["GET"])
def get_walkins():
    date = request.args.get("date", datetime.now().strftime("%Y-%m-%d"))
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM walkins WHERE date=%s ORDER BY id DESC", (date,))
    rows = rows_to_list(cur.fetchall(), cur)
    cur.execute("SELECT COALESCE(SUM(amount), 0) FROM walkins WHERE date=%s", (date,))
    total = cur.fetchone()[0]
    cur.close()
    conn.close()
    return jsonify({"walkins": rows, "total": round(float(total), 2), "date": date})

@app.route("/walkins", methods=["POST"])
def add_walkin():
    data = request.json or {}
    err = require_fields(data, "name", "amount")
    if err:
        return error(err)
    try:
        amount = float(data["amount"])
        if amount <= 0:
            return error("Amount must be greater than 0.")
    except (ValueError, TypeError):
        return error("Invalid amount.")
    date = data.get("date") or datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO walkins (name, amount, note, date) VALUES (%s, %s, %s, %s)",
        (data["name"].strip(), amount, data.get("note", "").strip(), date)
    )
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"message": "Walk-in recorded"}), 201

@app.route("/walkins/<int:walkin_id>", methods=["DELETE"])
def delete_walkin(walkin_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM walkins WHERE id=%s", (walkin_id,))
    conn.commit()
    rows_affected = cur.rowcount
    cur.close()
    conn.close()
    if rows_affected == 0:
        return error("Walk-in not found.", 404)
    return jsonify({"message": "Deleted"})

# ── Stats ─────────────────────────────────────────────────────────────────────

@app.route("/stats", methods=["GET"])
def stats():
    today = datetime.now().strftime("%Y-%m-%d")
    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM members")
    total = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM members WHERE expiration_date >= CURRENT_DATE")
    active = cur.fetchone()[0]

    cur.execute("SELECT COALESCE(SUM(price - (price * discount / 100)), 0) FROM members")
    revenue = cur.fetchone()[0]

    cur.execute(
        "SELECT COUNT(*) FROM members WHERE TO_CHAR(start_date::date, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')"
    )
    new_this_month = cur.fetchone()[0]

    cur.execute("SELECT COALESCE(SUM(amount), 0) FROM walkins WHERE date=%s", (today,))
    walkin_today = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM walkins WHERE date=%s", (today,))
    walkin_count = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM attendance WHERE date=%s AND time_out IS NULL", (today,))
    members_in_gym = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM attendance WHERE date=%s", (today,))
    visits_today = cur.fetchone()[0]

    cur.close()
    conn.close()

    return jsonify({
        "total_members": total,
        "active_members": active,
        "revenue": round(float(revenue), 2),
        "new_this_month": new_this_month,
        "walkin_revenue_today": round(float(walkin_today), 2),
        "walkin_count_today": walkin_count,
        "members_in_gym": members_in_gym,
        "visits_today": visits_today
    })

# ── Expiring ──────────────────────────────────────────────────────────────────

@app.route("/expiring", methods=["GET"])
def expiring():
    try:
        days = max(1, min(int(request.args.get("days", 7)), 365))
    except ValueError:
        days = 7
    today = datetime.now().strftime("%Y-%m-%d")
    future = (datetime.now() + timedelta(days=days)).strftime("%Y-%m-%d")
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM members WHERE expiration_date <= %s AND expiration_date >= %s ORDER BY expiration_date ASC",
        (future, today)
    )
    expiring_soon = rows_to_list(cur.fetchall(), cur)
    cur.execute(
        "SELECT * FROM members WHERE expiration_date < %s ORDER BY expiration_date DESC",
        (today,)
    )
    expired = rows_to_list(cur.fetchall(), cur)
    cur.close()
    conn.close()
    return jsonify({"expiring_soon": expiring_soon, "expired": expired})

# ── Export CSV ────────────────────────────────────────────────────────────────

@app.route("/export/csv", methods=["GET"])
def export_csv():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM members ORDER BY id DESC")
    rows = rows_to_list(cur.fetchall(), cur)
    cur.close()
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Name", "Email", "Phone", "Plan", "Months",
                     "Price", "Discount%", "Net Price", "Start Date", "Expiration Date", "Status"])
    today_str = datetime.now().strftime("%Y-%m-%d")
    for r in rows:
        net = r["price"] - (r["price"] * r["discount"] / 100)
        exp = r["expiration_date"]
        if exp < today_str:
            status = "Expired"
        elif (datetime.strptime(exp, "%Y-%m-%d") - datetime.now()).days <= 7:
            status = "Expiring Soon"
        else:
            status = "Active"
        writer.writerow([r["id"], r["name"], r["email"], r["phone"], r["plan"],
                         r["months"], r["price"], r["discount"], round(net, 2),
                         r["start_date"], r["expiration_date"], status])
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=members.csv"
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    return response

# ── Monthly Excel Report ──────────────────────────────────────────────────────

@app.route("/report/excel", methods=["GET"])
def report_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return error("openpyxl not installed.", 500)

    year = request.args.get("year", datetime.now().strftime("%Y"))
    month = request.args.get("month", datetime.now().strftime("%m")).zfill(2)
    month_str = f"{year}-{month}"
    try:
        month_name = datetime.strptime(month_str, "%Y-%m").strftime("%B %Y")
    except ValueError:
        return error("Invalid year/month.")

    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT * FROM members WHERE TO_CHAR(start_date::date,'YYYY-MM')=%s", (month_str,))
    new_members = rows_to_list(cur.fetchall(), cur)

    cur.execute("SELECT * FROM members ORDER BY name ASC")
    all_members = rows_to_list(cur.fetchall(), cur)

    cur.execute("SELECT * FROM members WHERE expiration_date >= CURRENT_DATE")
    active_members = rows_to_list(cur.fetchall(), cur)

    cur.execute("SELECT * FROM walkins WHERE TO_CHAR(date::date,'YYYY-MM')=%s ORDER BY date ASC", (month_str,))
    walkins = rows_to_list(cur.fetchall(), cur)

    cur.execute("""
        SELECT a.*, m.plan FROM attendance a
        LEFT JOIN members m ON a.member_id = m.id
        WHERE TO_CHAR(a.date::date,'YYYY-MM') = %s
        ORDER BY a.date ASC, a.time_in ASC
    """, (month_str,))
    attendance = rows_to_list(cur.fetchall(), cur)
    cur.close()
    conn.close()

    member_revenue = sum((r["price"] - r["price"] * r["discount"] / 100) for r in new_members)
    walkin_revenue = sum(r["amount"] for r in walkins)
    total_revenue = member_revenue + walkin_revenue

    DARK = "1a1a1a"; YELLOW = "E8FF00"; HEADER_BG = "2a2a2a"
    WHITE = "F0F0F0"; GREEN = "00E676"; MUTED = "888888"; ROW_ALT = "1f1f1f"
    thin = Side(style="thin", color="2a2a2a")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    def style_header_row(ws, row, cols, bg=HEADER_BG, fg=YELLOW):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True, color=fg, name="Arial", size=10)
            cell.fill = PatternFill("solid", start_color=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

    def style_data_row(ws, row, cols, alt=False):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(color=WHITE, name="Arial", size=9)
            cell.fill = PatternFill("solid", start_color=ROW_ALT if alt else DARK)
            cell.alignment = Alignment(vertical="center")
            cell.border = border

    def fill_bg(ws, max_row, max_col):
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.fill = PatternFill("solid", start_color=DARK)

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def sheet_title(ws, row, cols, text):
        ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
        c = ws[f"A{row}"]
        c.value = text
        c.font = Font(bold=True, color=YELLOW, name="Arial", size=14)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", start_color=DARK)
        ws.row_dimensions[row].height = 28

    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False
    fill_bg(ws1, 60, 10)
    ws1.merge_cells("A1:H1")
    c = ws1["A1"]
    c.value = "LOYD'S FITNESS GYM"
    c.font = Font(bold=True, color=YELLOW, name="Arial", size=22)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", start_color=DARK)
    ws1.row_dimensions[1].height = 40
    ws1.merge_cells("A2:H2")
    c = ws1["A2"]
    c.value = f"Monthly Report — {month_name}"
    c.font = Font(color=WHITE, name="Arial", size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill("solid", start_color=DARK)
    ws1.row_dimensions[2].height = 24

    kpis = [
        ("TOTAL MEMBERS", len(all_members), WHITE),
        ("ACTIVE MEMBERS", len(active_members), GREEN),
        ("NEW THIS MONTH", len(new_members), YELLOW),
        ("WALK-INS", len(walkins), "00B0FF"),
        ("ATTENDANCE LOGS", len(attendance), "FF9100"),
        ("MEMBER REVENUE", f"₱{member_revenue:,.2f}", GREEN),
        ("WALKIN REVENUE", f"₱{walkin_revenue:,.2f}", YELLOW),
        ("TOTAL REVENUE", f"₱{total_revenue:,.2f}", "FF4D00"),
    ]
    for col, (label, value, color) in enumerate(kpis, 1):
        lc = ws1.cell(row=5, column=col, value=label)
        lc.font = Font(color=MUTED, name="Arial", size=8, bold=True)
        lc.fill = PatternFill("solid", start_color=HEADER_BG)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = border
        vc = ws1.cell(row=6, column=col, value=value)
        vc.font = Font(bold=True, color=color, name="Arial", size=14)
        vc.fill = PatternFill("solid", start_color=HEADER_BG)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.border = border
    set_col_widths(ws1, [4, 20, 22, 14, 12, 8, 12, 10, 12, 12])

    ws2 = wb.create_sheet("All Members")
    ws2.sheet_view.showGridLines = False
    fill_bg(ws2, len(all_members) + 10, 11)
    sheet_title(ws2, 1, 11, f"ALL MEMBERS — {month_name}")
    h2 = ["#", "Name", "Email", "Phone", "Plan", "Months", "Net Price (₱)", "Discount %", "Start Date", "Expiration", "Status"]
    for i, h in enumerate(h2, 1):
        ws2.cell(row=2, column=i, value=h)
    style_header_row(ws2, 2, len(h2))
    today_str = datetime.now().strftime("%Y-%m-%d")
    for idx, m in enumerate(all_members):
        r = 3 + idx
        net = m["price"] - (m["price"] * m["discount"] / 100)
        exp = m["expiration_date"]
        days_left = (datetime.strptime(exp, "%Y-%m-%d") - datetime.now()).days
        if exp < today_str: status, color = "Expired", "FF1744"
        elif days_left <= 7: status, color = "Expiring Soon", "FFB300"
        else: status, color = "Active", GREEN
        for c, val in enumerate([idx+1, m["name"], m["email"], m["phone"], m["plan"],
                                  m["months"], round(net, 2), f'{m["discount"]}%',
                                  m["start_date"], exp, status], 1):
            ws2.cell(row=r, column=c, value=val)
        style_data_row(ws2, r, len(h2), alt=idx % 2 == 1)
        ws2.cell(row=r, column=11).font = Font(color=color, name="Arial", size=9, bold=True)
    set_col_widths(ws2, [4, 20, 22, 14, 12, 8, 14, 10, 12, 12, 14])

    ws3 = wb.create_sheet("Walk-ins")
    ws3.sheet_view.showGridLines = False
    fill_bg(ws3, len(walkins) + 10, 5)
    sheet_title(ws3, 1, 5, f"WALK-IN REVENUE — {month_name}")
    h3 = ["#", "Name", "Amount (₱)", "Note", "Date"]
    for i, h in enumerate(h3, 1):
        ws3.cell(row=2, column=i, value=h)
    style_header_row(ws3, 2, len(h3))
    for idx, w in enumerate(walkins):
        r = 3 + idx
        for c, val in enumerate([idx+1, w["name"], w["amount"], w["note"] or "—", w["date"]], 1):
            ws3.cell(row=r, column=c, value=val)
        style_data_row(ws3, r, len(h3), alt=idx % 2 == 1)
    set_col_widths(ws3, [4, 22, 14, 24, 14])

    ws4 = wb.create_sheet("Attendance")
    ws4.sheet_view.showGridLines = False
    fill_bg(ws4, len(attendance) + 10, 7)
    sheet_title(ws4, 1, 7, f"ATTENDANCE LOG — {month_name}")
    h4 = ["#", "Member", "Plan", "Date", "Time In", "Time Out", "Status"]
    for i, h in enumerate(h4, 1):
        ws4.cell(row=2, column=i, value=h)
    style_header_row(ws4, 2, len(h4))
    for idx, a in enumerate(attendance):
        r = 3 + idx
        status = "Inside" if not a["time_out"] else "Left"
        for c, val in enumerate([idx+1, a["member_name"], a.get("plan") or "—",
                                  a["date"], a["time_in"], a["time_out"] or "—", status], 1):
            ws4.cell(row=r, column=c, value=val)
        style_data_row(ws4, r, len(h4), alt=idx % 2 == 1)
        ws4.cell(row=r, column=7).font = Font(color=GREEN if status == "Inside" else "FF1744", name="Arial", size=9, bold=True)
    set_col_widths(ws4, [4, 22, 14, 12, 12, 12, 10])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"LoydsGym_Report_{month_name.replace(' ', '_')}.xlsx"
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
